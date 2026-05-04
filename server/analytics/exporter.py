"""
Módulo de exportación para pyme-ledger-ai.
Genera reportes en CSV (formato Excel-friendly), XLSX y PDF.
"""
import os
import csv
import json
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
from io import StringIO, BytesIO
from sqlalchemy.orm import Session
import pandas as pd

from models import Documento, Empresa

# Intentar importar reportlab para PDF
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

# Intentar importar openpyxl para Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExportEngine:
    """Motor de exportación de reportes."""

    def __init__(self, db: Session, empresa_id: str, data_dir: str = "data"):
        self.db = db
        self.empresa_id = empresa_id
        self.data_dir = Path(data_dir) / "exports"
        self.data_dir.mkdir(parents=True, exist_ok=True)

    def _get_docs(self, period_days: int = 30):
        """Obtiene documentos del período."""
        from datetime import timedelta
        start_date = datetime.utcnow() - timedelta(days=period_days)
        return self.db.query(Documento).filter(
            Documento.empresa_id == self.empresa_id,
            Documento.fecha_creacion >= start_date
        ).order_by(Documento.fecha_emision.desc()).all()

    def _get_empresa(self):
        """Obtiene la empresa."""
        return self.db.query(Empresa).filter(Empresa.id == self.empresa_id).first()

    def _build_rows(self, docs) -> List[Dict]:
        """Construye las filas de datos para exportación."""
        rows = []
        for doc in docs:
            rows.append({
                "Fecha Emisión": doc.fecha_emision.strftime("%d/%m/%Y") if doc.fecha_emision else "",
                "Tipo Documento": doc.tipo_documento.value if doc.tipo_documento else "",
                "Folio / N°": doc.folio or "",
                "Proveedor": doc.proveedor or "",
                "RUT / RUC": doc.rut_proveedor or "",
                "Descripción": (doc.texto_extraido or "")[:100].replace("\n", " ").strip(),
                "Categoría": doc.categoria.nombre if doc.categoria else "Sin asignar",
                "Centro de Costo": doc.centro_costo.nombre if doc.centro_costo else "—",
                "Moneda": doc.moneda or "CLP",
                "Monto Neto": round(doc.monto_neto, 2) if doc.monto_neto else 0.0,
                "IVA": round(doc.iva, 2) if doc.iva else 0.0,
                "Monto Total": round(doc.monto_total, 2) if doc.monto_total else 0.0,
                "Confianza IA": round(doc.confianza_clasificacion * 100, 0) if doc.confianza_clasificacion else 0,
                "Estado": doc.estado_revision.value if doc.estado_revision else "pendiente",
                "Alertas": self._format_exceptions(doc.excepciones),
                "Fecha Procesamiento": doc.fecha_creacion.strftime("%d/%m/%Y %H:%M") if doc.fecha_creacion else ""
            })
        return rows

    def export_to_csv(self, period_days: int = 30, filename: str = None) -> Dict:
        """
        Exporta documentos a CSV con formato Excel-friendly (separador ;, BOM UTF-8).
        """
        docs = self._get_docs(period_days)

        if not filename:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"gastos_{timestamp}.csv"

        file_path = self.data_dir / filename
        rows = self._build_rows(docs)

        if rows:
            df = pd.DataFrame(rows)
            # Usar ; como separador para que Excel lo abra correctamente en español
            df.to_csv(file_path, index=False, encoding="utf-8-sig", sep=";", decimal=",")
        else:
            with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=[
                    "Fecha Emisión", "Tipo Documento", "Folio / N°", "Proveedor",
                    "RUT / RUC", "Descripción", "Categoría", "Centro de Costo",
                    "Moneda", "Monto Neto", "IVA", "Monto Total",
                    "Confianza IA", "Estado", "Alertas", "Fecha Procesamiento"
                ], delimiter=";")
                writer.writeheader()

        return {
            "tipo": "CSV",
            "archivo": str(file_path),
            "filename": filename,
            "documentos": len(docs),
            "tamaño_bytes": file_path.stat().st_size if file_path.exists() else 0,
            "fecha_generacion": datetime.utcnow().isoformat()
        }

    def export_to_xlsx(self, period_days: int = 30, filename: str = None,
                       analytics_data: Dict = None) -> Dict:
        """
        Exporta documentos a Excel (.xlsx) con formato profesional:
        - Hoja 1: Detalle de gastos con formato de tabla
        - Hoja 2: Resumen por categoría
        - Hoja 3: Resumen por proveedor
        """
        if not HAS_OPENPYXL:
            # Fallback a CSV
            return self.export_to_csv(period_days, filename)

        docs = self._get_docs(period_days)
        empresa = self._get_empresa()

        if not filename:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"gastos_{timestamp}.xlsx"

        file_path = self.data_dir / filename
        rows = self._build_rows(docs)

        wb = openpyxl.Workbook()

        # ── Estilos ──
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="0D3DA6", end_color="0D3DA6", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_font = Font(name="Calibri", size=10)
        data_align = Alignment(vertical="center", wrap_text=False)
        money_format = '#,##0.00'
        pct_format = '0"%"'
        thin_border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0")
        )
        alt_fill = PatternFill(start_color="F2F6FF", end_color="F2F6FF", fill_type="solid")
        total_font = Font(name="Calibri", bold=True, size=11)
        total_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

        # ── Hoja 1: Detalle de Gastos ──
        ws1 = wb.active
        ws1.title = "Detalle de Gastos"
        ws1.sheet_properties.tabColor = "0D3DA6"

        # Título
        empresa_name = empresa.nombre_fantasia or empresa.razon_social if empresa else "Empresa"
        ws1.merge_cells("A1:P1")
        title_cell = ws1["A1"]
        title_cell.value = f"Reporte de Gastos — {empresa_name}"
        title_cell.font = Font(name="Calibri", bold=True, size=16, color="0D3DA6")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 35

        # Subtítulo
        ws1.merge_cells("A2:P2")
        sub_cell = ws1["A2"]
        sub_cell.value = f"Período: últimos {period_days} días | Generado: {datetime.utcnow().strftime('%d/%m/%Y %H:%M')} | Documentos: {len(docs)}"
        sub_cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
        sub_cell.alignment = Alignment(horizontal="center")
        ws1.row_dimensions[2].height = 22

        # Headers (fila 4)
        headers = [
            "Fecha Emisión", "Tipo Documento", "Folio / N°", "Proveedor",
            "RUT / RUC", "Descripción", "Categoría", "Centro de Costo",
            "Moneda", "Monto Neto", "IVA", "Monto Total",
            "Confianza IA", "Estado", "Alertas", "Fecha Procesamiento"
        ]
        col_widths = [14, 16, 12, 28, 16, 30, 18, 16, 8, 14, 14, 14, 12, 12, 25, 18]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws1.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws1.column_dimensions[get_column_letter(col_idx)].width = width

        ws1.row_dimensions[4].height = 28

        # Datos
        money_cols = {10, 11, 12}  # Monto Neto, IVA, Monto Total (1-indexed)
        pct_col = 13  # Confianza IA

        for row_idx, row_data in enumerate(rows, 5):
            for col_idx, header in enumerate(headers, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

                # Formato numérico para montos
                if col_idx in money_cols:
                    cell.number_format = money_format
                elif col_idx == pct_col:
                    cell.number_format = pct_format

                # Filas alternas
                if (row_idx - 5) % 2 == 1:
                    cell.fill = alt_fill

        # Fila de totales
        if rows:
            total_row = 5 + len(rows)
            ws1.cell(row=total_row, column=1, value="TOTALES").font = total_font
            for col_idx in range(1, len(headers) + 1):
                cell = ws1.cell(row=total_row, column=col_idx)
                cell.fill = total_fill
                cell.border = thin_border
                cell.font = total_font

            # Sumar montos
            for col_idx in money_cols:
                cell = ws1.cell(row=total_row, column=col_idx)
                col_letter = get_column_letter(col_idx)
                cell.value = f"=SUM({col_letter}5:{col_letter}{total_row - 1})"
                cell.number_format = money_format

            ws1.cell(row=total_row, column=9, value=f"{len(rows)} docs").font = total_font

        # Freeze panes
        ws1.freeze_panes = "A5"
        # Auto-filter
        ws1.auto_filter.ref = f"A4:P{4 + len(rows)}"

        # ── Hoja 2: Resumen por Categoría ──
        ws2 = wb.create_sheet("Resumen por Categoría")
        ws2.sheet_properties.tabColor = "3DAE2B"

        # Agrupar por categoría
        cat_summary = {}
        for row_data in rows:
            cat = row_data.get("Categoría", "Sin asignar")
            if cat not in cat_summary:
                cat_summary[cat] = {"total": 0.0, "neto": 0.0, "iva": 0.0, "cantidad": 0}
            cat_summary[cat]["total"] += row_data.get("Monto Total", 0)
            cat_summary[cat]["neto"] += row_data.get("Monto Neto", 0)
            cat_summary[cat]["iva"] += row_data.get("IVA", 0)
            cat_summary[cat]["cantidad"] += 1

        # Ordenar por total descendente
        sorted_cats = sorted(cat_summary.items(), key=lambda x: x[1]["total"], reverse=True)
        grand_total = sum(v["total"] for v in cat_summary.values()) or 1

        ws2.merge_cells("A1:F1")
        ws2["A1"].value = "Resumen por Categoría"
        ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="0D3DA6")
        ws2["A1"].alignment = Alignment(horizontal="center")

        cat_headers = ["Categoría", "Cantidad", "Monto Neto", "IVA", "Monto Total", "% del Total"]
        cat_widths = [25, 12, 16, 16, 16, 14]
        for col_idx, (h, w) in enumerate(zip(cat_headers, cat_widths), 1):
            cell = ws2.cell(row=3, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color="3DAE2B", end_color="3DAE2B", fill_type="solid")
            cell.alignment = header_align
            cell.border = thin_border
            ws2.column_dimensions[get_column_letter(col_idx)].width = w

        for r_idx, (cat_name, vals) in enumerate(sorted_cats, 4):
            ws2.cell(row=r_idx, column=1, value=cat_name).font = data_font
            ws2.cell(row=r_idx, column=2, value=vals["cantidad"]).font = data_font
            c_neto = ws2.cell(row=r_idx, column=3, value=round(vals["neto"], 2))
            c_neto.font = data_font
            c_neto.number_format = money_format
            c_iva = ws2.cell(row=r_idx, column=4, value=round(vals["iva"], 2))
            c_iva.font = data_font
            c_iva.number_format = money_format
            c_total = ws2.cell(row=r_idx, column=5, value=round(vals["total"], 2))
            c_total.font = data_font
            c_total.number_format = money_format
            c_pct = ws2.cell(row=r_idx, column=6, value=round(vals["total"] / grand_total * 100, 1))
            c_pct.font = data_font
            c_pct.number_format = '0.0"%"'
            for c in range(1, 7):
                ws2.cell(row=r_idx, column=c).border = thin_border
                if (r_idx - 4) % 2 == 1:
                    ws2.cell(row=r_idx, column=c).fill = alt_fill

        ws2.freeze_panes = "A4"

        # ── Hoja 3: Resumen por Proveedor ──
        ws3 = wb.create_sheet("Resumen por Proveedor")
        ws3.sheet_properties.tabColor = "00B4D8"

        prov_summary = {}
        for row_data in rows:
            prov = row_data.get("Proveedor", "Desconocido") or "Desconocido"
            if prov not in prov_summary:
                prov_summary[prov] = {"total": 0.0, "cantidad": 0, "categorias": set()}
            prov_summary[prov]["total"] += row_data.get("Monto Total", 0)
            prov_summary[prov]["cantidad"] += 1
            prov_summary[prov]["categorias"].add(row_data.get("Categoría", ""))

        sorted_provs = sorted(prov_summary.items(), key=lambda x: x[1]["total"], reverse=True)

        ws3.merge_cells("A1:E1")
        ws3["A1"].value = "Resumen por Proveedor"
        ws3["A1"].font = Font(name="Calibri", bold=True, size=14, color="0D3DA6")
        ws3["A1"].alignment = Alignment(horizontal="center")

        prov_headers = ["Proveedor", "Cantidad Docs", "Monto Total", "% del Total", "Categorías"]
        prov_widths = [30, 14, 16, 14, 35]
        for col_idx, (h, w) in enumerate(zip(prov_headers, prov_widths), 1):
            cell = ws3.cell(row=3, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color="00B4D8", end_color="00B4D8", fill_type="solid")
            cell.alignment = header_align
            cell.border = thin_border
            ws3.column_dimensions[get_column_letter(col_idx)].width = w

        for r_idx, (prov_name, vals) in enumerate(sorted_provs[:50], 4):  # Top 50
            ws3.cell(row=r_idx, column=1, value=prov_name).font = data_font
            ws3.cell(row=r_idx, column=2, value=vals["cantidad"]).font = data_font
            c_total = ws3.cell(row=r_idx, column=3, value=round(vals["total"], 2))
            c_total.font = data_font
            c_total.number_format = money_format
            c_pct = ws3.cell(row=r_idx, column=4, value=round(vals["total"] / grand_total * 100, 1))
            c_pct.font = data_font
            c_pct.number_format = '0.0"%"'
            ws3.cell(row=r_idx, column=5, value=", ".join(sorted(vals["categorias"]))).font = data_font
            for c in range(1, 6):
                ws3.cell(row=r_idx, column=c).border = thin_border
                if (r_idx - 4) % 2 == 1:
                    ws3.cell(row=r_idx, column=c).fill = alt_fill

        ws3.freeze_panes = "A4"

        # Guardar
        wb.save(str(file_path))

        return {
            "tipo": "XLSX",
            "archivo": str(file_path),
            "filename": filename,
            "documentos": len(docs),
            "hojas": ["Detalle de Gastos", "Resumen por Categoría", "Resumen por Proveedor"],
            "tamaño_bytes": file_path.stat().st_size if file_path.exists() else 0,
            "fecha_generacion": datetime.utcnow().isoformat()
        }

    def export_to_pdf(self, analytics_data: Dict, filename: str = None) -> Dict:
        """Exporta reporte a PDF."""
        if not HAS_REPORTLAB:
            return {"error": "reportlab no está instalado", "sugerencia": "pip install reportlab"}

        if not filename:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"reporte_{timestamp}.pdf"

        file_path = self.data_dir / filename

        try:
            empresa = self._get_empresa()
            doc = SimpleDocTemplate(str(file_path), pagesize=letter)
            story = []
            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(
                'CustomTitle', parent=styles['Heading1'],
                fontSize=22, textColor=colors.HexColor('#0D3DA6'),
                spaceAfter=30, alignment=TA_CENTER
            )
            heading_style = ParagraphStyle(
                'CustomHeading', parent=styles['Heading2'],
                fontSize=13, textColor=colors.HexColor('#0D3DA6'),
                spaceAfter=12, spaceBefore=12
            )

            # Título
            empresa_name = empresa.nombre_fantasia or empresa.razon_social if empresa else "Empresa"
            story.append(Paragraph(f"Reporte de Gastos — {empresa_name}", title_style))
            story.append(Spacer(1, 0.3 * inch))

            # Info general
            story.append(Paragraph("Información General", heading_style))
            general_data = [
                ["Empresa:", empresa.razon_social if empresa else ""],
                ["RUT:", empresa.rut if empresa else ""],
                ["Período:", f"{analytics_data.get('periodo_dias', 30)} días"],
                ["Generado:", datetime.utcnow().strftime("%d/%m/%Y %H:%M")]
            ]
            general_table = Table(general_data, colWidths=[2 * inch, 4 * inch])
            general_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F0FE')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D0D0D0'))
            ]))
            story.append(general_table)
            story.append(Spacer(1, 0.3 * inch))

            # KPIs
            if analytics_data.get('kpis'):
                story.append(Paragraph("Indicadores Clave", heading_style))
                kpis = analytics_data['kpis']
                kpi_data = [
                    ["Métrica", "Valor"],
                    ["Total Gasto", f"${kpis.get('total_gasto', 0):,.2f}"],
                    ["Total Neto", f"${kpis.get('total_neto', 0):,.2f}"],
                    ["Total IVA", f"${kpis.get('total_iva', 0):,.2f}"],
                    ["Promedio por Documento", f"${kpis.get('promedio_documento', 0):,.2f}"],
                    ["Documentos", str(kpis.get('cantidad_documentos', 0))],
                    ["Pendientes", str(kpis.get('documentos_pendiente_revision', 0))],
                    ["Revisados", str(kpis.get('documentos_revisados', 0))]
                ]
                kpi_table = Table(kpi_data, colWidths=[3 * inch, 3 * inch])
                kpi_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D3DA6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D0D0D0')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F6FF')])
                ]))
                story.append(kpi_table)
                story.append(Spacer(1, 0.3 * inch))

            # Categorías
            if analytics_data.get('categorias'):
                story.append(Paragraph("Gastos por Categoría", heading_style))
                cat_data = [["Categoría", "Total", "%", "Docs"]]
                for cat in analytics_data['categorias'][:15]:
                    cat_data.append([
                        cat.get('categoria', ''),
                        f"${cat.get('total', 0):,.2f}",
                        f"{cat.get('porcentaje', 0):.1f}%",
                        str(cat.get('cantidad', 0))
                    ])
                cat_table = Table(cat_data, colWidths=[2.5 * inch, 1.5 * inch, 1 * inch, 1 * inch])
                cat_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3DAE2B')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D0D0D0')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F6FF')])
                ]))
                story.append(cat_table)

            doc.build(story)
            return {
                "tipo": "PDF",
                "archivo": str(file_path),
                "filename": filename,
                "tamaño_bytes": file_path.stat().st_size if file_path.exists() else 0,
                "fecha_generacion": datetime.utcnow().isoformat()
            }
        except Exception as e:
            return {"error": f"Error generando PDF: {str(e)}"}

    def export_accounting_ledger(self, period_days: int = 30, filename: str = None) -> Dict:
        """Exporta libro contable en formato estándar."""
        docs = self._get_docs(period_days)
        # Ordenar por fecha ascendente para el libro
        docs.sort(key=lambda d: d.fecha_emision or datetime.min)

        if not filename:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"libro_contable_{timestamp}.csv"

        file_path = self.data_dir / filename

        rows = []
        saldo = 0.0
        for doc in docs:
            monto = doc.monto_total or 0
            saldo -= monto
            rows.append({
                "Fecha": doc.fecha_emision.strftime("%d/%m/%Y") if doc.fecha_emision else "",
                "Descripción": f"{doc.tipo_documento.value if doc.tipo_documento else 'Doc'} — {doc.proveedor or 'Desconocido'}",
                "Cuenta": doc.categoria.codigo if doc.categoria and hasattr(doc.categoria, 'codigo') else "0000",
                "Categoría": doc.categoria.nombre if doc.categoria else "Sin asignar",
                "Debe": round(monto, 2),
                "Haber": 0.00,
                "Saldo": round(saldo, 2)
            })

        if rows:
            df = pd.DataFrame(rows)
            df.to_csv(file_path, index=False, encoding="utf-8-sig", sep=";", decimal=",")

        return {
            "tipo": "Libro Contable CSV",
            "archivo": str(file_path),
            "filename": filename,
            "documentos": len(docs),
            "saldo_final": f"{saldo:.2f}",
            "tamaño_bytes": file_path.stat().st_size if file_path.exists() else 0,
            "fecha_generacion": datetime.utcnow().isoformat()
        }

    def _format_exceptions(self, exceptions_json: str) -> str:
        """Formatea excepciones para exportación."""
        if not exceptions_json:
            return ""
        try:
            exceptions = json.loads(exceptions_json)
            if isinstance(exceptions, list):
                return "; ".join([e.get("message", "") for e in exceptions[:3]])
            return ""
        except Exception:
            return ""


# Instancia global
export_engine = None


def get_export_engine(db: Session, empresa_id: str, data_dir: str = "data") -> ExportEngine:
    """Obtiene una instancia del motor de exportación."""
    return ExportEngine(db, empresa_id, data_dir)
